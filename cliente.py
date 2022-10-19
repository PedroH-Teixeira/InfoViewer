import sys
import time
from PyQt5.QtWidgets import QMainWindow, QApplication
from designer.principal import *
from threading import Thread
from os import path, mkdir
from datetime import datetime
import openpyxl

class App(QMainWindow, Ui_TelaPrincipal):
    def __init__(self, parent=None):
        super().__init__(parent)
        super().setupUi(self)

        Thread(target=self.atualiza_ano, daemon=True).start()

        self.pushButton_Enviar.clicked.connect(self.setValorPlanilha)

        try:
            if not path.exists(f'{self.ano_atual}'):
                mkdir(f'{self.ano_atual}')

            if not path.exists(fr'{self.ano_atual}/dados.xlsx'):
                planilha = openpyxl.Workbook()
                aba_ativa = planilha.create_sheet('dados', 0)

                data = [
                    ['LOCAL', 'CONTAINER', '', 'STATUS', 'DATA'],
                    ['LOCAL', 'CONTAINER', '', 'STATUS', 'DATA'],
                    ['LOCAL', 'CONTAINER', '', 'STATUS', 'DATA'],
                    ['LOCAL', 'CONTAINER', '', 'STATUS', 'DATA'],
                    ['LOCAL', 'CONTAINER', '', 'STATUS', 'DATA'],
                    ['LOCAL', 'CONTAINER', '', 'STATUS', 'DATA'],
                    ['LOCAL', 'CONTAINER', '', 'STATUS', 'DATA'],
                    ['LOCAL', 'CONTAINER', '', 'STATUS', 'DATA'],
                    ['LOCAL', 'CONTAINER', '', 'STATUS', 'DATA'],
                    ['LOCAL', 'CONTAINER', '', 'STATUS', 'DATA'],
                ]

                # add column headings. NB. these must be strings
                aba_ativa.append(["Local", "Container", "Placa","Status","Data"])
                for row in data:
                    aba_ativa.append(row)

                tabela = openpyxl.worksheet.table.Table(displayName="dados", ref="A1:E11")

                # Add a default style with striped rows and banded columns
                estilo_tabela = openpyxl.worksheet.table.TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                                               showLastColumn=False, showRowStripes=True, showColumnStripes=True)
                tabela.tableStyleInfo = estilo_tabela

                # Cria aba e salva tabela
                aba_ativa.add_table(tabela)
                planilha.save(fr'{self.ano_atual}/dados.xlsx')
        except Exception as erro:
            print(erro)

    def setValorPlanilha(self):
        try:
            planilha = openpyxl.load_workbook(fr'{self.ano_atual}/dados.xlsx')
            aba_ativa = planilha.worksheets[0]

            valor1 = self.comboBox_Local.currentText()
            valor2 = f"{self.lineEdit_Container_1.text().upper()}-{self.lineEdit_Container_2.text().upper()}-" \
                     f"{self.lineEdit_Container_3.text().upper()}"
            valor3 = self.lineEdit_Placa.text().upper()
            valor4 = self.comboBox_Status.currentText()
            data_atual = datetime.now()
            valor5 = data_atual.strftime('%d/%m - %H:%M')

            if valor1 == '' or valor2 == '' or valor3 == '' or valor4 == '':
                self.label_Informativo.setStyleSheet('color: blue;')
                self.label_Informativo.setText('Todos os campos devem ser preenchidos.')

            else:
                linha = [valor1, valor2, valor3, valor4, valor5]

                aba_ativa.append(linha)
                ultima_linha = aba_ativa.max_row

                tabela = aba_ativa.tables['dados']
                tabela.ref = f"A1:E{ultima_linha}"

                planilha.save(fr'{self.ano_atual}/dados.xlsx')
                self.comboBox_Status.setCurrentText('')
                self.label_Informativo.setStyleSheet('color: green;')
                self.label_Informativo.setText(f'Informações enviadas.')
        except OSError as erro:
            self.label_Informativo.setStyleSheet('color: red;')
            self.label_Informativo.setText(f'{erro}')

    def atualiza_ano(self):
        while True:
            self.ano_atual = datetime.today().year
            time.sleep(86400)

if __name__ == '__main__':
    qt = QApplication(sys.argv)
    app = App()
    app.show()
    qt.exec_()