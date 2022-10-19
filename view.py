import sys
import time

from PyQt5.QtWidgets import QMainWindow, QApplication
from designer.viewer import *
from threading import Thread
from os import path
from datetime import datetime
import openpyxl
from pygame import mixer

class App(QMainWindow, Ui_View):
    def __init__(self):
        super().__init__()
        super().setupUi(self)

        self.ano_atual = datetime.today().year

        Thread(target=self.atualizar_planilha, daemon=True).start()

    def atualizar_planilha(self):

        if path.exists(fr'{self.ano_atual}/dados.xlsx'):

            beep =0

            while True:

                planilha = openpyxl.load_workbook(fr'{self.ano_atual}/dados.xlsx')
                aba_ativa = planilha.worksheets[0]
                total_linhas = aba_ativa.max_row

                if beep != total_linhas:
                    beep = total_linhas
                    try:
                        mixer.init()
                        sound = mixer.Sound(r'img/beep.wav')
                        sound.play()
                    except Exception as erro:
                        print(erro)

                valorLocal = aba_ativa.cell(total_linhas, 1).value
                valorContainer = aba_ativa.cell(total_linhas, 2).value
                valorPlaca = aba_ativa.cell(total_linhas, 3).value
                valorStatus = aba_ativa.cell(total_linhas, 4).value
                valorData = aba_ativa.cell(total_linhas, 5).value

                valorLocal2 = aba_ativa.cell(total_linhas - 1, 1).value
                valorContainer2 = aba_ativa.cell(total_linhas - 1, 2).value
                valorPlaca2 = aba_ativa.cell(total_linhas - 1, 3).value
                valorStatus2 = aba_ativa.cell(total_linhas- 1, 4).value
                valorData2 = aba_ativa.cell(total_linhas - 1, 5).value

                valorLocal3 = aba_ativa.cell(total_linhas - 2, 1).value
                valorContainer3 = aba_ativa.cell(total_linhas - 2, 2).value
                valorPlaca3 = aba_ativa.cell(total_linhas - 2, 3).value
                valorStatus3 = aba_ativa.cell(total_linhas - 2, 4).value
                valorData3 = aba_ativa.cell(total_linhas - 2, 5).value

                valorLocal4 = aba_ativa.cell(total_linhas - 3, 1).value
                valorContainer4 = aba_ativa.cell(total_linhas - 3, 2).value
                #valorPlaca4 = aba_ativa.cell(total_linhas - 3, 3).value
                valorStatus4 = aba_ativa.cell(total_linhas - 3, 4).value
                valorData4 = aba_ativa.cell(total_linhas - 3, 5).value

                valorLocal5 = aba_ativa.cell(total_linhas - 4, 1).value
                valorContainer5 = aba_ativa.cell(total_linhas - 4, 2).value
                #valorPlaca5 = aba_ativa.cell(total_linhas - 4, 3).value
                valorStatus5 = aba_ativa.cell(total_linhas - 4, 4).value
                valorData5 = aba_ativa.cell(total_linhas - 4, 5).value

                valorLocal6 = aba_ativa.cell(total_linhas - 5, 1).value
                valorContainer6 = aba_ativa.cell(total_linhas - 5, 2).value
                #valorPlaca6 = aba_ativa.cell(total_linhas - 5, 3).value
                valorStatus6 = aba_ativa.cell(total_linhas - 5, 4).value
                valorData6 = aba_ativa.cell(total_linhas - 5, 5).value

                valorLocal7 = aba_ativa.cell(total_linhas - 6, 1).value
                valorContainer7 = aba_ativa.cell(total_linhas - 6, 2).value
                #valorPlaca7 = aba_ativa.cell(total_linhas - 6, 3).value
                valorStatus7 = aba_ativa.cell(total_linhas - 6, 4).value
                valorData7 = aba_ativa.cell(total_linhas - 6, 5).value

                valorLocal8 = aba_ativa.cell(total_linhas - 7, 1).value
                valorContainer8 = aba_ativa.cell(total_linhas - 7, 2).value
                #valorPlaca8 = aba_ativa.cell(total_linhas - 7, 3).value
                valorStatus8 = aba_ativa.cell(total_linhas - 7, 4).value
                valorData8 = aba_ativa.cell(total_linhas - 7, 5).value

                valorLocal9 = aba_ativa.cell(total_linhas - 8, 1).value
                valorContainer9 = aba_ativa.cell(total_linhas - 8, 2).value
                #valorPlaca9 = aba_ativa.cell(total_linhas - 8, 3).value
                valorStatus9 = aba_ativa.cell(total_linhas - 8, 4).value
                valorData9 = aba_ativa.cell(total_linhas - 8, 5).value

                valorLocal10 = aba_ativa.cell(total_linhas - 9, 1).value
                valorContainer10 = aba_ativa.cell(total_linhas - 9, 2).value
                #valorPlaca10 = aba_ativa.cell(total_linhas - 9, 3).value
                valorStatus10 = aba_ativa.cell(total_linhas - 9, 4).value
                valorData10 = aba_ativa.cell(total_linhas - 9, 5).value

                # ===================================================

                self.label_valorLocal.setText(valorLocal)
                self.label_valorContainer.setText(valorContainer)
                self.label_valorPlaca.setText(valorPlaca)
                self.label_status.setText(valorStatus)
                self.label_data.setText(valorData)

                self.label_valorLocal2.setText(valorLocal2)
                self.label_valorContainer2.setText(valorContainer2)
                self.label_valorPlaca_2.setText(valorPlaca2)
                self.label_status2.setText(valorStatus2)
                self.label_data2.setText(valorData2)

                self.label_valorLocal3.setText(valorLocal3)
                self.label_valorContainer3.setText(valorContainer3)
                self.label_valorPlaca_3.setText(valorPlaca3)
                self.label_status3.setText(valorStatus3)
                self.label_data3.setText(valorData3)

                self.label_valorLocal4.setText(valorLocal4)
                self.label_valorContainer4.setText(valorContainer4)
                self.label_status4.setText(valorStatus4)
                self.label_data4.setText(valorData4)

                self.label_valorLocal5.setText(valorLocal5)
                self.label_valorContainer5.setText(valorContainer5)
                self.label_status5.setText(valorStatus5)
                self.label_data5.setText(valorData5)

                self.label_valorLocal6.setText(valorLocal6)
                self.label_valorContainer6.setText(valorContainer6)
                self.label_status6.setText(valorStatus6)
                self.label_data6.setText(valorData6)

                self.label_valorLocal7.setText(valorLocal7)
                self.label_valorContainer7.setText(valorContainer7)
                self.label_status7.setText(valorStatus7)
                self.label_data7.setText(valorData7)

                self.label_valorLocal8.setText(valorLocal8)
                self.label_valorContainer8.setText(valorContainer8)
                self.label_status8.setText(valorStatus8)
                self.label_data8.setText(valorData8)

                self.label_valorLocal9.setText(valorLocal9)
                self.label_valorContainer9.setText(valorContainer9)
                self.label_status9.setText(valorStatus9)
                self.label_data9.setText(valorData9)

                self.label_valorLocal10.setText(valorLocal10)
                self.label_valorContainer10.setText(valorContainer10)
                self.label_status10.setText(valorStatus10)
                self.label_data10.setText(valorData10)
                time.sleep(10)

if __name__ == '__main__':
    qt = QApplication(sys.argv)
    app = App()
    app.showFullScreen()
    qt.exec_()